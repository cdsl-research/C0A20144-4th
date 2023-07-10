import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import math
import xlrd
import socket
import subprocess
import pickle
import os

host = "送り先のIP"
port = 1234

# Excelファイルのパスとシート名
file_path = 'control-file.xlsx'
sheet_name = 'control'

# セル範囲を指定
start_row = 2
end_row = 5
column = 'A'

# 指定のディレクトリからファイル一覧を取得
directory = '送りたいファイルがあるディレクトリのパス'  # 適切なディレクトリパスに修正する
files_in_directory = os.listdir(directory)

# 選択肢のリストを作成
choices = []
files = []  # ファイル情報を保持するリスト
for file_name in files_in_directory:
    choices.append(file_name)
    file_info = {
        'file_name': file_name,
        'priority': None,  # ファイルの優先度を保持する変数
        'project': None,  # ファイルが属するプロジェクト名を保持する変数
    }
    files.append(file_info)

# 選択肢を表示
print("送信するファイルを選択してください:")
for i, choice in enumerate(choices, start=1):
    print(f"{i}. {choice}")

# 選択肢の番号を入力
selection = input("番号を入力してください: ")

# 選択されたファイルを送信する処理を実行
if selection.isdigit() and int(selection) in range(1, len(choices) + 1):
    selected_file = choices[int(selection) - 1]
    selected_info = files[int(selection) - 1]  # 選択されたファイルの情報を取得

    selected_file_path = os.path.join(directory, selected_file)  # 適切なファイルパスに修正する

    # エクセルファイルを読み込む
    workbook = load_workbook(file_path, data_only=True)

    # シートを選択する
    sheet = workbook[sheet_name]

    # プロジェクト名のリストを作成
    project_names = []
    for row in range(start_row, end_row + 1):
        project_cell = sheet['A{}'.format(row)]
        project_name = project_cell.value
        if project_name:
            project_names.append(project_name)

    # 選択肢を表示
    print("ファイルが属するプロジェクトを選択してください:")
    for i, project_name in enumerate(project_names, start=1):
        print(f"{i}. {project_name}")

    # 選択肢の番号を入力
    project_selection = input("番号を入力してください: ")

    # 選択されたプロジェクトを設定
    if project_selection.isdigit() and int(project_selection) in range(1, len(project_names) + 1):
        selected_project = project_names[int(project_selection) - 1]
        selected_info['project'] = selected_project

        workbook = load_workbook(file_path, data_only=True)

        # シートを選択する
        sheet = workbook[sheet_name]

        # セルの値を取得する
        start_dates = []
        nouki_dates = []
        work_date = None
        project_dates = []

        for row in range(start_row, end_row + 1):
            start_cell = sheet['B{}'.format(row)]
            nouki_cell = sheet['C{}'.format(row)]
            project_cell = sheet['D{}'.format(row)]

            start_date = start_cell.value
            nouki_date = nouki_cell.value
            project_date = project_cell.value

            start_date = datetime(*xlrd.xldate_as_tuple(start_date, 0)).date()
            nouki_date = datetime(*xlrd.xldate_as_tuple(nouki_date, 0)).date()

            start_dates.append(start_date)
            nouki_dates.append(nouki_date)
            project_dates.append(project_date)

        work_date_cell = sheet['E2']
        work_date_value = work_date_cell.value
        work_date = datetime(*xlrd.xldate_as_tuple(work_date_value, 0)).date()

        # 日数として取得する(作業日数の算出)
        work_days = [nouki - start for start, nouki in zip(start_dates, nouki_dates)]

        # 日数として取得する(残り日数の算出)
        remain_days = [nouki - work_date for nouki in nouki_dates]

        priorities = []

        # プロジェクトの現在の進捗
        for project_date, work_day, remain_day in zip(project_dates, work_days, remain_days):
            day = work_day.days
            remain = remain_day.days
            priority = math.floor((day - remain) / day * 100)
            ans = priority - project_date
            priorities.append(ans)

        selected_info = files[int(selection) - 1]  # 選択されたファイルの情報を取得
        selected_info['priority'] = priorities[int(project_selection) - 1]  # 優先度として対応する値を設定

        # ファイルの情報と紐付けられた数字を出力
        print("ファイルの情報:")
        for file_info in files:
            print(f"ファイル名: {file_info['file_name']}")
            print(f"優先度: {file_info['priority']}")

        # データのシリアライズ
        serialized_data = pickle.dumps(selected_info['priority'])

        # Socketを使用してデータを送信
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.connect((host, port))
        s.send(serialized_data)
        s.close()

        # ここにファイル送信の処理を追加してください
        try:
            source_file = selected_file_path
            destination_file = '送り先のパス：例~@IP:パス'
            command = ['scp', source_file, destination_file]
            process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            stdout, stderr = process.communicate()

        except subprocess.CalledProcessError as e:
            print(f"ファイルの送信中にエラーが発生しました: {e}")

    else:
        print("無効な選択です。")

else:
    print("無効な選択です。")

